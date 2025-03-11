import streamlit as st
import requests
import os
import time

import PyPDF2
from PyPDF2 import PdfReader, PdfWriter
import io
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

API_URL = os.environ.get("BACKEND_API_URL")

def wait_for_backend(message="Waiting for backend to start...", retry_interval=2, max_retries=30):
    """
    Display a waiting message while attempting to connect to the backend.
    
    Args:
        message (str): The message to display.
        retry_interval (int): Number of seconds to wait between retries.
        max_retries (int): Maximum number of retries before giving up.
        
    Returns:
        bool: True if the backend is available, False if max retries were exceeded.
    """
    api_url = os.environ.get("BACKEND_API_URL")
    
    with st.spinner(message):
        for _ in range(max_retries):
            try:
                # Try to connect to any endpoint
                response = requests.get(f"{api_url}/api/ms_auth/get_auth_url", timeout=2)
                if response.status_code == 200:
                    return True
            except requests.RequestException:
                pass
            
            time.sleep(retry_interval)
    
    return False

class BaseAPIClient:
    """Base class to handle connection with the FastAPI with retry mechanism."""
    def __init__(self, api_url=None):
        self.api_url = api_url or os.environ.get("BACKEND_API_URL")

    def _request(self, method, endpoint, **kwargs):
        """Helper method to send requests to the API with retry logic."""
        try:
            response = method(f"{self.api_url}{endpoint}", **kwargs)
            response.raise_for_status()
            return response.json()
        except requests.RequestException as e:
            # Check if it's a connection error (backend not started yet)
            if isinstance(e, (requests.ConnectionError, requests.Timeout)):
                if wait_for_backend():
                    # Try again after backend is available
                    try:
                        response = method(f"{self.api_url}{endpoint}", **kwargs)
                        response.raise_for_status()
                        return response.json()
                    except:
                        return {"error": "Error processing request after backend started"}
                else:
                    return {"error": "Backend service is not available. Please try again later."}
            return {"error": f"Could not connect to the server: {e}"}

    def _get(self, endpoint):
        """Helper method to send GET requests to the API."""
        return self._request(requests.get, endpoint)

    def _post(self, endpoint, json=None):
        """Helper method to send POST requests to the API."""
        return self._request(requests.post, endpoint, json=json)
    
def log_function():
    logout_url = requests.get(f"{API_URL}/api/ms_auth/get_logout_url").json()
    aida_url = requests.get(f"{API_URL}/api/ms_auth/get_aida_url").json()

    st.markdown(f'''
        <div style="text-align: center;">
            <div style="display: inline-block; width: 48%; text-align: center;">
                <a href="{logout_url}" target="_self" style="text-decoration:none;">
                    <button style="padding:5px 10px; font-size:15px; width:100%; 
                            background-color: #f44336; color: white; 
                            border: none; border-radius: 4px;">
                        🔒 Logout
                    </button>
                </a>
         
        </div>
    ''', unsafe_allow_html=True)
    


def detect_scanned_pdf(pdf_bytes: bytes) -> bool:
    """
    Checks if the total number of characters in a PDF file provided as bytes is 1000 or more.
    Returns True if the PDF is likely to be scanned or needs OCR based on the character count.

    Args:
        pdf_bytes (bytes): PDF file in bytes format.

    Returns:
        bool: True if the total number of characters is 1000 or more, indicating the need for OCR,
              False otherwise.
    """
    total_characters: int = 0  # Initialize total_characters counter

    try:
        # Create a BytesIO object to handle the byte input
        pdf_stream = BytesIO(pdf_bytes)
        pdf_reader = PyPDF2.PdfReader(pdf_stream)  # Create a PDF reader object

        # Iterate through each page in the PDF
        for page in pdf_reader.pages:
            # Extract text from the page and count characters
            text = page.extract_text()
            if text:  # Check if text extraction was successful
                total_characters += len(text)

            # Return True if the total characters exceed 1000
            if total_characters >= 10:
                return True

        # Return False if the total characters are less than 1000
        return False
    
    except Exception as e:
        # Print the error message and return False if an error occurs
        print(f"Error processing PDF: {e}")
        return False
 
   
def get_table_dimensions(content):
    """
    Obtiene las dimensiones de una tabla markdown.
    
    Args:
        content (str): Contenido de la tabla en formato Markdown.
    
    Returns:
        tuple: (número de filas, número de columnas)
    """
    rows = [row for row in content.splitlines() if "|" in row and "---" not in row]
    if not rows:
        return 0, 0
    columns = len(rows[0].split("|")[1:-1])
    return len(rows), columns

def process_table(content, sheet, start_row, role, fill=None):
    """
    Procesa una tabla representada como texto y la inserta en una hoja Excel.
    
    Args:
        content (str): Contenido de la tabla en formato Markdown o similar.
        sheet (openpyxl.Worksheet): Hoja donde se insertará la tabla.
        start_row (int): Fila inicial para insertar la tabla.
        role (str): Rol asociado a la tabla (User o Assistant).
        fill (PatternFill, optional): Color de fondo para las celdas.
    """
    rows = [row.split("|")[1:-1] for row in content.splitlines() if "|" in row and "---" not in row]
    for i, row in enumerate(rows):
        current_row = start_row + i
        for j, cell in enumerate(row, 2):  # Comenzar en columna B para dejar la columna A para el rol
            cell_value = cell.strip().replace("&lt;br&gt;", "\n")
            excel_cell = sheet.cell(row=current_row, column=j, value=cell_value)
            excel_cell.alignment = Alignment(wrap_text=True, vertical="center")
            excel_cell.font = Font(name="Calibri", size=11)
            excel_cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            if fill:
                excel_cell.fill = fill
        # Asegurar que el rol esté presente en la columna A
        if i == 0:  # Solo agregar el rol en la primera fila de la tabla
            role_cell = sheet.cell(row=current_row, column=1, value=role)
            role_cell.alignment = Alignment(horizontal="center", vertical="center")
            role_cell.font = Font(name="Calibri", size=11)

def add_text_cell(sheet, row_num, role, content, fill):
    """
    Añade una celda de texto con el formato apropiado.
    
    Args:
        sheet (openpyxl.Worksheet): Hoja de trabajo
        row_num (int): Número de fila
        role (str): Rol (User o Assistant)
        content (str): Contenido del texto
        fill (PatternFill): Relleno para la celda
    """
    cell_font = Font(name="Calibri", size=11)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    
    sheet.cell(row=row_num, column=1, value=role).font = cell_font
    sheet.cell(row=row_num, column=1).alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(row=row_num, column=1).fill = fill

    sheet.cell(row=row_num, column=2, value=content.replace("&lt;br&gt;", "\n")).font = cell_font
    sheet.cell(row=row_num, column=2).alignment = cell_alignment
    sheet.cell(row=row_num, column=2).fill = fill

def apply_fill_to_range(sheet, start_row, end_row, start_col, end_col, fill):
    """
    Aplica un color de fondo a un rango de celdas.
    
    Args:
        sheet (openpyxl.Worksheet): Hoja de trabajo
        start_row (int): Fila inicial
        end_row (int): Fila final
        start_col (int): Columna inicial
        end_col (int): Columna final
        fill (PatternFill): Color de fondo a aplicar
    """
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            if not cell.fill or cell.fill.start_color.rgb == "00000000":
                cell.fill = fill


def create_chat_excel(conversation):
    """
    Crea un archivo Excel con formato atractivo para la conversación, manejando tablas y texto.
    Devuelve los bytes del archivo Excel para ser descargado con Streamlit.
    
    Args:
        conversation (list): Lista de diccionarios con 'role' y 'content'.
    
    Returns:
        bytes: Bytes del archivo Excel generado.
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "RALFR3 - AIDA 3.0"

    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 100

    # Configurar encabezados
    headers = ["Rol", "Contenido"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    row_fill_user = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    row_fill_assistant = PatternFill(start_color="DFF2BF", end_color="DFF2BF", fill_type="solid")

    row_num = 2
    for msg in conversation:
        role = msg['role'].capitalize()
        content = msg['content']
        fill = row_fill_user if role == 'User' else row_fill_assistant

        if "|" in content and "---" in content:
            # Separar el contenido en texto antes, tabla y texto después
            parts = content.split("\n\n")
            current_row = row_num
            start_row = current_row  # Guardamos la fila inicial
            max_columns = 2  # Mínimo número de columnas (Rol y Contenido)
            
            for part in parts:
                if "|" in part and "---" in part:
                    # Es una tabla
                    rows, columns = get_table_dimensions(part)
                    max_columns = max(max_columns, columns + 1)  # +1 por la columna del rol
                    process_table(part, sheet, current_row, role, fill)
                    current_row += rows
                elif part.strip():
                    # Es texto normal
                    add_text_cell(sheet, current_row, role, part.strip(), fill)
                    current_row += 1
            
            # Aplicar color de fondo a toda la región de la respuesta
            if role == "Assistant":
                apply_fill_to_range(sheet, start_row, current_row - 1, 1, max_columns, fill)
            
            row_num = current_row
        else:
            # Contenido normal sin tabla
            add_text_cell(sheet, row_num, role, content, fill)
            row_num += 1

    # Guardar el archivo en un objeto BytesIO
    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)  # Reposicionar al inicio del archivo para su lectura
    return excel_bytes.getvalue()